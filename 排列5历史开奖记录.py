# Copyright (c) 2019 Zhang_Ray
#
# -*- coding:utf-8 -*-
# @Script: 排列5历史走势分析.py
# @作者: Zhang_Ray
# @Email: 1228506574@qq.com
# @Create At: 2019-04-30 23:36:40
# @最后编辑者: Zhang_Ray
# @最后编辑时间: 2019-05-10 13:41:05
# @文件说明: This is Spyder for lottery_pls, then analysis and deside to buy.

import numpy as np
import requests
import xlrd
import xlsxwriter
import xlwt
import pandas as pd
from os import walk
from lxml import etree
from openpyxl import Workbook, worksheet
from xlutils.copy import copy
from xlutils.filter import XLWTWriter

#爬取该网站200个网页的开奖数据
page = 1
pl1 = []
tl1 = []
tp = []
tl = []

#获取开奖时间并保持
while page <= 200:
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'}
    url = 'http://www.lottery.gov.cn/historykj/history_%d.jspx?_ltype=pls' % page
    r = requests.get(url,headers=header)
    html = etree.HTML(r.text)
    time = html.xpath('//div[@class="result"]/table/tbody/tr/td[11]/text()')
    #nums = html.xpath('//div[@class="result"]/table/tbody/tr/td[2]/text()')
    time = list(time)

    #建立.xlsx文件
    #根据page保存文件名为同样长度的文件
    if page < 10:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3\\ForTest00%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        #将time list的str 分别 保存到.xlsx文件
        for i in range(len(time)-1):
            worksheet.write(i,0,time[i])
    
    if 10 <= page < 100:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3\\ForTest0%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        for i in range(len(time)-1):
            worksheet.write(i,0,time[i])
    
    if page >= 100:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3\\ForTest%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        for i in range(len(time)-1):
            worksheet.write(i,0,time[i])
    
    workbook.close()
    print('已保存第%d页开奖时间' % page)
    page += 1

#获取开奖结果并保存
page = 1
while page <= 200:
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'}
    url = 'http://www.lottery.gov.cn/historykj/history_%d.jspx?_ltype=pls' % page
    r = requests.get(url,headers=header)
    html = etree.HTML(r.text)
    #time = html.xpath('//div[@class="result"]/table/tbody/tr/td[11]/text()')
    nums = html.xpath('//div[@class="result"]/table/tbody/tr/td[2]/text()')
    nums = list(nums)
    #建立.xlsx文件
    #根据page保存文件名为同样长度的文件
    if page < 10:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3nums\\00%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        for i in range(len(nums)-1):
            worksheet.write(i,0,nums[i])
    
    if 10 <= page < 100:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3nums\\0%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        for i in range(len(nums)-1):
            worksheet.write(i,0,nums[i])
    
    if page >= 100:
        workbook = xlsxwriter.Workbook("E:\\AllHomeWorkHere\\pl3nums\\%d.xlsx" % page)
        worksheet = workbook.add_worksheet()
        for i in range(len(nums)-1):
            worksheet.write(i,0,nums[i])
    
    workbook.close()
    print('已保存第%d页开奖结果' % page)
    page += 1

print('200页的开奖结果及时间保存完毕')


'''
    i = 1
    values = []
    while i < page:
        try:
            f=xlrd.open_workbook("C:\\USERS\\12285\\DESKTOP\\ForTest%d.xlsx" % i)
            print('ook')
            sheet=f.sheets()
            rn = sheet.nrows

        except Exception as e:
            print("打开文件错误："+e)
        print('ok')
        i += 1
        print(sheet)
'''

'''
    try:
        i = 0
        l3 = []
        while i <= len(time):
            timen = time[i]
            l3.append(timen)
            i += 1
    except:
        pass
    workbook = xlsxwriter.Workbook("C:\\USERS\\12285\\DESKTOP\\ForTest.xlsx")
    worksheet = workbook.add_worksheet()
    for i in range(len(l3)-1):
        worksheet.write(i,0,l3[i])
'''

'''
    try:
        for i in time:
            time1 = list3[0:i]
            print(time1)
            i += 1
    except:
        workbook = Workbook()
        booksheet = workbook.active     #获取当前活跃的sheet,默认是第一个sheet
        for i in range(1,len(time)):
            booksheet.cell(1,i).value = time[1]   #这个方法索引从1开始
            i += 1
            booksheet.append([nums,time])
            workbook.save("C:\\USERS\\12285\\DESKTOP\\ForTest.xls")
'''

'''
    tl1.append(time)
    pl1.append(nums)
    #pl1.append(r)
'''

'''
    for i in range(len(tl1)):
        tp.append(tl1[i] + pl1[i])
'''

'''
    try:
        workbook = Workbook()
        booksheet = workbook.active     #获取当前活跃的sheet,默认是第一个sheet
        for i in range(1,len(tl1)):
            booksheet.cell(1,i).value = tl1[i]   #这个方法索引从1开始
            i += 1
        booksheet.append([pl1,tl1])
        workbook.save("C:\\USERS\\12285\\DESKTOP\\ForTest.xls")
    except:
        pass

    try:
        xlsfile = 'ForTest.xls'
        book = xlrd.open_workbook(xlsfile)
        sheet_name = book.sheet_names()
        print(sheet_name)
    except:
        pass    

    print(tl1)
    print(pl1)
    print(tp)
'''
